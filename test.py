




import math
import matplotlib
# matplotlib.get_backend()
matplotlib.use('TkAgg')
# matplotlib.use('WXagg')
import matplotlib.pyplot as plt
import numpy as np
import matplotlib
import os
import sys
import torch




import openpyxl as op

def write():
    num_list = [1,2,3,4,5,6]
    L = 59
    p = np.array([0.2, 0.2, 0.2, 0.2, 0.2, 0.1, 0.1, 0.02, 0.01, 0.01, 0.01])
    p = p/sum(p)
    num_list = np.random.choice([100, 99, 98, 97, 96, 95, 94, 93, 92, 91, 90], size = L, replace=True, p = p)


    bg = op.load_workbook(r"hahaha.xlsx")      	# 应先将excel文件放入到工作目录下
    sheet = bg["Sheet1"]                          		 	# “Sheet1”表示将数据写入到excel文件的sheet1下
    for i in range(1, len(num_list)+1):
        sheet.cell(i , 1, num_list[i - 1])					# sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
    bg.save("hahaha.xlsx")            			# 对文件进行保存

write()









































































































































