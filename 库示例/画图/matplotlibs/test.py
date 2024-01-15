#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun 13 14:16:14 2023

@author: jack
"""

import matplotlib
# matplotlib.get_backend()
matplotlib.use('TkAgg')
# matplotlib.use('WXagg')
import matplotlib.pyplot as plt
import numpy as np
import matplotlib
from matplotlib.font_manager import FontProperties
from pylab import tick_params
import copy
from matplotlib.pyplot import MultipleLocator

# matplotlib.use('Agg')
# matplotlib.rcParams['text.usetex'] = True
filepath2 = '/home/jack/snap/'
font = FontProperties(fname="/usr/share/fonts/truetype/msttcorefonts/Times_New_Roman.ttf", size=14)

fontpath = "/usr/share/fonts/truetype/windows/"
# fname =  "/usr/share/fonts/truetype/arphic/SimSun.ttf",
font = FontProperties(fname=fontpath+"simsun.ttf", size=22)

fontpath1 = "/usr/share/fonts/truetype/msttcorefonts/"
fonte = FontProperties(fname=fontpath1+"Times_New_Roman.ttf", size=22)
fonte1 = FontProperties(fname=fontpath1+"Times_New_Roman.ttf", size=24)

fontpath2 = "/usr/share/fonts/truetype/NerdFonts/"
font1 = FontProperties(fname=fontpath2+"Caskaydia Cove ExtraLight Nerd Font Complete.otf", size=20)
font1 = FontProperties(fname=fontpath2+"Caskaydia Cove Light Nerd Font Complete Mono.otf", size=20)
font1 = FontProperties(fname=fontpath2+"Caskaydia Cove SemiLight Nerd Font Complete.otf", size=20)
font1 = FontProperties(fname=fontpath2+"Caskaydia Cove Regular Nerd Font Complete Mono.otf", size=20)



##==========================================  2 ===================================================
import numpy as np
# loadtxt()中的dtype参数默认设置为float
# 这里设置为str字符串便于显示
a = np.loadtxt('tmp.txt', )




res = a[:, 0:-3].mean(axis = 1) * 0.2  + a[:, -2]*0.2 + a[:, -1]*0.6


res = np.ceil(res)


import openpyxl as op

num_list = [1,2,3,4,5,6]
L = 59
# p = np.array([0.2, 0.2, 0.2, 0.2, 0.2, 0.1, 0.1, 0.02, 0.01, 0.01, 0.01])
# p = p/sum(p)
# num_list = np.random.choice([100, 99, 98, 97, 96, 95, 94, 93, 92, 91, 90], size = L, replace=True, p = p)


bg = op.load_workbook("/home/jack/snap/res.xlsx")      	# 应先将excel文件放入到工作目录下
sheet = bg["Sheet1"]                          		 	# “Sheet1”表示将数据写入到excel文件的sheet1下
for i in range(1, len(res)+1):
    sheet.cell(i , 1, res[i - 1])					# sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
bg.save("/home/jack/snap/res.xlsx")            			# 对文件进行保存















